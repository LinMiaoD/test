"""
CSV / XLSX 文件差异比对工具 - 带图形界面
支持：CSV vs CSV、XLSX vs CSV、CSV vs XLSX、XLSX vs XLSX
"""

import csv
import os
from datetime import datetime
from itertools import zip_longest
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ─────────────────────────── 文件读取 ───────────────────────────

def detect_encoding(filepath):
    with open(filepath, "rb") as f:
        raw = f.read(4)
    if raw[:3] == b"\xef\xbb\xbf":
        return "utf-8-sig"
    try:
        with open(filepath, encoding="utf-8") as f:
            f.read()
        return "utf-8"
    except UnicodeDecodeError:
        return "gbk"


def read_csv(filepath):
    enc = detect_encoding(filepath)
    with open(filepath, newline="", encoding=enc) as f:
        reader = csv.DictReader(f)
        headers = list(reader.fieldnames or [])
        rows = [dict(row) for row in reader]
    return headers, rows


def read_xlsx(filepath, sheet_index=0):
    """读取 xlsx 文件，返回 (headers, rows)，rows 为 list[dict]"""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    if sheet_index >= len(sheet_names):
        sheet_index = 0
    ws = wb[sheet_names[sheet_index]]

    rows_raw = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows_raw:
        return [], []

    headers = [str(c) if c is not None else "" for c in rows_raw[0]]
    rows = []
    for raw in rows_raw[1:]:
        row = {}
        for h, v in zip(headers, raw):
            row[h] = str(v) if v is not None else ""
        rows.append(row)
    return headers, rows


def get_sheets(filepath):
    """返回 xlsx 的工作表列表；csv 返回 ['Sheet1']"""
    ext = os.path.splitext(filepath)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xls"):
        if not HAS_OPENPYXL:
            return ["Sheet1"]
        wb = openpyxl.load_workbook(filepath, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names
    return ["Sheet1"]


def read_file(filepath, sheet_index=0):
    """统一入口：根据扩展名自动选择读取方式"""
    ext = os.path.splitext(filepath)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xls"):
        if not HAS_OPENPYXL:
            raise RuntimeError("读取 xlsx 需要安装 openpyxl 库：pip install openpyxl")
        return read_xlsx(filepath, sheet_index)
    else:
        return read_csv(filepath)


# ─────────────────────────── 比对核心 ───────────────────────────

def normalize(val):
    v = (val or "").strip()
    if v == "":
        return ""
    try:
        f = float(v)
        return str(int(f)) if f == int(f) else str(f)
    except ValueError:
        return v


def compare_files(file1, file2, sheet1=0, sheet2=0):
    lines = []

    def log(msg=""):
        lines.append(msg)

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log(f"比对时间: {ts}")
    log(f"文件1: {file1}")
    log(f"文件2: {file2}")

    # 显示工作表信息（仅 xlsx）
    ext1 = os.path.splitext(file1)[1].lower()
    ext2 = os.path.splitext(file2)[1].lower()
    if ext1 in (".xlsx", ".xlsm", ".xls"):
        sheets1 = get_sheets(file1)
        log(f"文件1 工作表: {sheets1[sheet1] if sheet1 < len(sheets1) else sheet1}")
    if ext2 in (".xlsx", ".xlsm", ".xls"):
        sheets2 = get_sheets(file2)
        log(f"文件2 工作表: {sheets2[sheet2] if sheet2 < len(sheets2) else sheet2}")

    log("=" * 60)

    h1, rows1 = read_file(file1, sheet1)
    h2, rows2 = read_file(file2, sheet2)

    only1 = [c for c in h1 if c not in h2]
    only2 = [c for c in h2 if c not in h1]
    if only1:
        log(f"[列] 仅文件1存在: {only1}")
    if only2:
        log(f"[列] 仅文件2存在: {only2}")

    if len(rows1) != len(rows2):
        log(f"[行数] 文件1={len(rows1)}行  文件2={len(rows2)}行  相差{abs(len(rows1) - len(rows2))}行")

    common = [c for c in h1 if c in h2]
    diff_count = 0

    for i, (r1, r2) in enumerate(zip_longest(rows1, rows2), start=1):
        if r1 is None:
            row_str = "  ".join(f"{c}={r2.get(c, '')}" for c in common)
            log(f"[行{i}] + 文件2新增: {row_str}")
            diff_count += 1
        elif r2 is None:
            row_str = "  ".join(f"{c}={r1.get(c, '')}" for c in common)
            log(f"[行{i}] - 文件1删除: {row_str}")
            diff_count += 1
        else:
            diffs = []
            for c in common:
                v1 = normalize(r1.get(c))
                v2 = normalize(r2.get(c))
                if v1 != v2:
                    raw1 = (r1.get(c) or "").strip()
                    raw2 = (r2.get(c) or "").strip()
                    diffs.append(f'{c}: "{raw1}"->"{raw2}"')
            if diffs:
                log(f"[行{i}] ~ " + " | ".join(diffs))
                diff_count += 1

    log("=" * 60)
    log(f"共 {diff_count} 处差异" if diff_count else "无差异")

    return "\n".join(lines), diff_count


# ─────────────────────────── GUI ───────────────────────────

FILETYPES = [("支持的文件", "*.csv *.xlsx *.xlsm"), ("CSV 文件", "*.csv"),
             ("Excel 文件", "*.xlsx *.xlsm"), ("所有文件", "*.*")]

XLSX_EXTS = (".xlsx", ".xlsm", ".xls")


class SheetSelector(tk.Frame):
    """文件路径 + 工作表下拉，封装为一行组件"""
    def __init__(self, master, label, bg, **kw):
        super().__init__(master, bg=bg, **kw)
        self.bg = bg
        self._sheets = ["Sheet1"]

        tk.Label(self, text=label, font=("Microsoft YaHei", 9, "bold"),
                 bg=bg, width=12, anchor="w").grid(row=0, column=0, pady=4)

        self.path_var = tk.StringVar()
        self.path_var.trace_add("write", self._on_path_change)
        tk.Entry(self, textvariable=self.path_var, font=("Consolas", 9),
                 width=48, relief="solid", bd=1).grid(row=0, column=1, padx=6)

        tk.Button(self, text="浏览…", command=self._browse,
                  font=("Microsoft YaHei", 9), bg="#3498DB", fg="white",
                  relief="flat", padx=10, cursor="hand2").grid(row=0, column=2, padx=4)

        # 工作表行（初始隐藏）
        self._sheet_row = tk.Frame(self, bg=bg)
        self._sheet_row.grid(row=1, column=0, columnspan=3, sticky="w", pady=(0, 2))
        tk.Label(self._sheet_row, text="工作表:", font=("Microsoft YaHei", 9),
                 bg=bg, fg="#7F8C8D").pack(side="left")
        self.sheet_var = tk.StringVar()
        self._combo = ttk.Combobox(self._sheet_row, textvariable=self.sheet_var,
                                   state="readonly", width=30,
                                   font=("Microsoft YaHei", 9))
        self._combo.pack(side="left", padx=6)
        self._sheet_row.grid_remove()

    def _browse(self):
        path = filedialog.askopenfilename(title="选择文件", filetypes=FILETYPES)
        if path:
            self.path_var.set(path)

    def _on_path_change(self, *_):
        path = self.path_var.get().strip()
        ext = os.path.splitext(path)[1].lower()
        if ext in XLSX_EXTS and os.path.isfile(path):
            try:
                sheets = get_sheets(path)
                self._combo["values"] = sheets
                self.sheet_var.set(sheets[0])
                self._sheet_row.grid()
                return
            except Exception:
                pass
        self._sheet_row.grid_remove()

    def get_path(self):
        return self.path_var.get().strip()

    def get_sheet_index(self):
        path = self.get_path()
        ext = os.path.splitext(path)[1].lower()
        if ext in XLSX_EXTS:
            name = self.sheet_var.get()
            try:
                sheets = get_sheets(path)
                return sheets.index(name) if name in sheets else 0
            except Exception:
                return 0
        return 0


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("CSV / XLSX 差异比对工具")
        self.geometry("780x660")
        self.resizable(True, True)
        self.configure(bg="#F5F6FA")
        self._result_text = ""
        self._build_ui()

    def _build_ui(self):
        tk.Label(self, text="CSV / XLSX 差异比对工具",
                 font=("Microsoft YaHei", 16, "bold"),
                 bg="#F5F6FA", fg="#2C3E50").pack(pady=(18, 2))
        tk.Label(self, text="支持 CSV ↔ CSV、XLSX ↔ CSV、XLSX ↔ XLSX 任意组合比对",
                 font=("Microsoft YaHei", 9), bg="#F5F6FA", fg="#7F8C8D").pack()

        if not HAS_OPENPYXL:
            tk.Label(self, text="⚠  未检测到 openpyxl，xlsx 文件支持不可用（pip install openpyxl）",
                     font=("Microsoft YaHei", 9), bg="#FFF3CD", fg="#856404",
                     relief="flat", padx=8).pack(fill="x", padx=20, pady=(6, 0))

        frame = tk.LabelFrame(self, text=" 选择文件 ", font=("Microsoft YaHei", 10),
                              bg="#F5F6FA", fg="#2C3E50", padx=12, pady=10)
        frame.pack(fill="x", padx=20, pady=10)

        self.sel1 = SheetSelector(frame, "文件1 (基准)", bg="#F5F6FA")
        self.sel1.pack(fill="x", pady=2)
        ttk.Separator(frame, orient="horizontal").pack(fill="x", pady=4)
        self.sel2 = SheetSelector(frame, "文件2 (对比)", bg="#F5F6FA")
        self.sel2.pack(fill="x", pady=2)

        btn_frame = tk.Frame(self, bg="#F5F6FA")
        btn_frame.pack(pady=6)

        self.run_btn = tk.Button(btn_frame, text="▶  开始比对", command=self._run,
                                 font=("Microsoft YaHei", 11, "bold"),
                                 bg="#27AE60", fg="white", relief="flat",
                                 padx=24, pady=8, cursor="hand2")
        self.run_btn.pack(side="left", padx=8)

        tk.Button(btn_frame, text="💾  保存日志", command=self._save_log,
                  font=("Microsoft YaHei", 10), bg="#8E44AD", fg="white",
                  relief="flat", padx=16, pady=8, cursor="hand2").pack(side="left", padx=8)

        tk.Button(btn_frame, text="🗑  清空", command=self._clear,
                  font=("Microsoft YaHei", 10), bg="#95A5A6", fg="white",
                  relief="flat", padx=16, pady=8, cursor="hand2").pack(side="left", padx=8)

        out_frame = tk.LabelFrame(self, text=" 比对结果 ", font=("Microsoft YaHei", 10),
                                  bg="#F5F6FA", fg="#2C3E50", padx=8, pady=8)
        out_frame.pack(fill="both", expand=True, padx=20, pady=(2, 14))

        self.text = tk.Text(out_frame, font=("Consolas", 9), wrap="none",
                            bg="#1E1E2E", fg="#CDD6F4", insertbackground="white",
                            relief="flat", state="disabled")
        vsb = ttk.Scrollbar(out_frame, orient="vertical", command=self.text.yview)
        hsb = ttk.Scrollbar(out_frame, orient="horizontal", command=self.text.xview)
        self.text.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self.text.pack(fill="both", expand=True)

        self.text.tag_config("add",    foreground="#A6E3A1")
        self.text.tag_config("del",    foreground="#F38BA8")
        self.text.tag_config("mod",    foreground="#FAB387")
        self.text.tag_config("warn",   foreground="#F9E2AF")
        self.text.tag_config("header", foreground="#89DCEB",
                             font=("Consolas", 9, "bold"))

        self.status_var = tk.StringVar(value="就绪")
        tk.Label(self, textvariable=self.status_var, font=("Microsoft YaHei", 9),
                 bg="#F5F6FA", fg="#7F8C8D", anchor="w").pack(fill="x", padx=20, pady=(0, 6))

    def _clear(self):
        self.sel1.path_var.set("")
        self.sel2.path_var.set("")
        self._set_text("")
        self._result_text = ""
        self.status_var.set("已清空")

    def _run(self):
        f1 = self.sel1.get_path()
        f2 = self.sel2.get_path()

        if not f1 and not f2:
            messagebox.showwarning("提示", "请先选择或输入两个文件路径！")
            return
        if not f1:
            messagebox.showwarning("提示", "文件1 路径为空，请填写！")
            return
        if not f2:
            messagebox.showwarning("提示", "文件2 路径为空，请填写！")
            return
        if not os.path.isfile(f1):
            messagebox.showerror("错误", f"文件1 不存在：\n{f1}")
            return
        if not os.path.isfile(f2):
            messagebox.showerror("错误", f"文件2 不存在：\n{f2}")
            return

        s1 = self.sel1.get_sheet_index()
        s2 = self.sel2.get_sheet_index()

        self.run_btn.config(state="disabled", text="比对中…")
        self.status_var.set("正在比对，请稍候…")
        threading.Thread(target=self._do_compare,
                         args=(f1, f2, s1, s2), daemon=True).start()

    def _do_compare(self, f1, f2, s1, s2):
        try:
            result, diff_count = compare_files(f1, f2, s1, s2)
            self._result_text = result
            self.after(0, lambda: self._show_result(result, diff_count))
        except Exception as e:
            self.after(0, lambda: messagebox.showerror("比对出错", str(e)))
            self.after(0, lambda: self.run_btn.config(state="normal",
                                                       text="▶  开始比对"))

    def _show_result(self, result, diff_count):
        self._set_text(result)
        if diff_count == 0:
            self.status_var.set("✅ 比对完成，无差异")
        else:
            self.status_var.set(f"⚠️  比对完成，共发现 {diff_count} 处差异")
        self.run_btn.config(state="normal", text="▶  开始比对")

    def _set_text(self, content):
        self.text.config(state="normal")
        self.text.delete("1.0", "end")
        for line in content.splitlines():
            if line.startswith("[行") and "] +" in line:
                tag = "add"
            elif line.startswith("[行") and "] -" in line:
                tag = "del"
            elif line.startswith("[行") and "] ~" in line:
                tag = "mod"
            elif line.startswith("[列") or line.startswith("[行数"):
                tag = "warn"
            elif line.startswith("="):
                tag = "header"
            else:
                tag = None
            self.text.insert("end", line + "\n", tag if tag else "")
        self.text.config(state="disabled")
        self.text.see("1.0")

    def _save_log(self):
        if not self._result_text:
            messagebox.showinfo("提示", "还没有比对结果，请先运行比对！")
            return
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = filedialog.asksaveasfilename(
            title="保存日志",
            initialfile=f"diff_{ts}.log",
            defaultextension=".log",
            filetypes=[("日志文件", "*.log"), ("文本文件", "*.txt"), ("所有文件", "*.*")]
        )
        if path:
            with open(path, "w", encoding="utf-8") as f:
                f.write(self._result_text)
            messagebox.showinfo("保存成功", f"日志已保存至：\n{path}")
            self.status_var.set(f"日志已保存: {path}")


if __name__ == "__main__":
    app = App()
    app.mainloop()
