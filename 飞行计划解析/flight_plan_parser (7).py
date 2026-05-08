"""
飞行日计划表解析器
==================
Windows CMD 下直接把Excel文件拖进窗口即可解析。

用法:
    1. 双击运行本脚本，按提示拖入文件
    2. 或在CMD中: python flight_plan_parser.py "C:\路径\计划表.xlsx"
"""

import re
import sys
import os
from dataclasses import dataclass
from typing import Optional
from collections import OrderedDict

import openpyxl


# ============================================================
# 配置常量
# ============================================================
START_ROW = 9
HOUR_TIME_ROW = 7
MIN_TIME_ROW = 8
CODE_CIPHER_START_ROW = 9
CODE_CIPHER_GROUPS = [(3, 4), (8, 9), (13, 14)]
COURSE_PATTERN = re.compile(r'^(LLNR|FM|LLN)[\w-]*$')
PERSON_PATTERN = re.compile(r'^([\u4e00-\u9fff])(\d{3,4})$')


@dataclass
class FlightSortie:
    course: str
    start_time: str
    end_time: str
    front_seat_daizi: str
    front_seat_daihao: str
    rear_seat_daizi: Optional[str]
    rear_seat_daihao: Optional[str]
    is_solo: bool
    row: int
    start_col: int
    end_col: int


class FlightPlanParser:
    def __init__(self, filepath: str, sheet_name: str = None):
        self.filepath = filepath
        self.wb = openpyxl.load_workbook(filepath)
        self.ws = self.wb[sheet_name] if sheet_name else self.wb.active
        self.data_start_col = None
        self.data_end_col = None
        self.time_map = {}
        self.code_cipher_map = OrderedDict()

    def parse(self) -> list:
        self._find_boundaries()
        self._build_time_map()
        self._build_code_cipher_map()
        return self._extract_sorties()

    def _find_boundaries(self):
        for col in range(1, self.ws.max_column + 1):
            val = self.ws.cell(row=HOUR_TIME_ROW, column=col).value
            if val == '飞机':
                self.data_start_col = col
            elif val == '备注':
                self.data_end_col = col
        if self.data_start_col is None:
            raise ValueError("未找到'飞机'列")
        if self.data_end_col is None:
            raise ValueError("未找到'备注'列")

    def _build_time_map(self):
        # 取飞机列之前最近的小时
        current_hour = None
        for col in range(self.data_start_col - 1, 0, -1):
            h = self.ws.cell(row=HOUR_TIME_ROW, column=col).value
            if h and isinstance(h, str) and ':' in h:
                current_hour = int(h.split(':')[0].strip())
                break
        # 每列=5分钟; row8有分钟值的列是基准，下一列=基准+5分钟
        last_base_min = None
        for col in range(self.data_start_col + 1, self.data_end_col):
            h = self.ws.cell(row=HOUR_TIME_ROW, column=col).value
            m = self.ws.cell(row=MIN_TIME_ROW, column=col).value
            if h and isinstance(h, str) and ':' in h:
                current_hour = int(h.split(':')[0].strip())
            if m is not None and current_hour is not None:
                base_min = int(m)
                self.time_map[col] = f"{current_hour}:{str(base_min).zfill(2)}"
                last_base_min = base_min
            elif last_base_min is not None and current_hour is not None:
                # 无分钟值的列 = 上一个基准分钟 + 5
                next_min = last_base_min + 5
                hr = current_hour
                if next_min >= 60:
                    next_min -= 60
                    hr += 1
                self.time_map[col] = f"{hr}:{str(next_min).zfill(2)}"
                last_base_min = None  # 已用完，等下一个基准

    def _build_code_cipher_map(self):
        for row in range(CODE_CIPHER_START_ROW, self.ws.max_row + 1):
            for dz_col, dh_col in CODE_CIPHER_GROUPS:
                dz = self.ws.cell(row=row, column=dz_col).value
                dh = self.ws.cell(row=row, column=dh_col).value
                if dz is not None and dh is not None:
                    self.code_cipher_map[str(dz).strip()] = str(dh).strip()

    def _extract_sorties(self) -> list:
        results = []
        max_data_row = self._find_max_data_row()
        for row in range(START_ROW, max_data_row + 1):
            for col in range(self.data_start_col + 1, self.data_end_col):
                val = self.ws.cell(row=row, column=col).value
                if val is None:
                    continue
                val_str = str(val).strip()
                if not COURSE_PATTERN.match(val_str):
                    continue
                sortie = self._parse_single_sortie(val_str, row, col)
                if sortie:
                    results.append(sortie)
        return results

    def _parse_single_sortie(self, course, row, col):
        start_time = self.time_map.get(col, "未知")
        end_col = None
        front_dz = front_dh = None

        for scan_col in range(col + 1, self.data_end_col):
            v = self.ws.cell(row=row, column=scan_col).value
            if v is None:
                continue
            match = PERSON_PATTERN.match(str(v).strip())
            if not match:
                continue
            dz, dh = match.group(1), match.group(2)
            if dz in self.code_cipher_map and self.code_cipher_map[dz] == dh:
                front_dz, front_dh = dz, dh
                end_col = scan_col
                break
        if end_col is None:
            return None

        end_time = "未知"
        for c in range(end_col - 1, col - 1, -1):
            if c in self.time_map:
                end_time = self.time_map[c]
                break

        rear_dz = rear_dh = None
        is_solo = True
        next_val = self.ws.cell(row=row + 1, column=end_col).value
        if next_val is not None:
            match = PERSON_PATTERN.match(str(next_val).strip())
            if match:
                r_dz, r_dh = match.group(1), match.group(2)
                if r_dz in self.code_cipher_map and self.code_cipher_map[r_dz] == r_dh:
                    rear_dz, rear_dh = r_dz, r_dh
                    is_solo = False

        return FlightSortie(
            course=course, start_time=start_time, end_time=end_time,
            front_seat_daizi=front_dz, front_seat_daihao=front_dh,
            rear_seat_daizi=rear_dz, rear_seat_daihao=rear_dh,
            is_solo=is_solo, row=row, start_col=col, end_col=end_col,
        )

    def _find_max_data_row(self):
        empty_count = 0
        last_row = START_ROW
        for row in range(START_ROW, self.ws.max_row + 1):
            has_data = any(
                self.ws.cell(row=row, column=c).value is not None
                for c in range(self.data_start_col + 1, self.data_end_col)
            )
            if has_data:
                last_row = row
                empty_count = 0
            else:
                empty_count += 1
                if empty_count >= 3:
                    break
        return last_row


def col_to_letter(col_num):
    result = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result


def SEP(char="=", length=60):
    print(char * length)


def print_results(parser, sorties, filepath):
    print()
    SEP()
    print(f"  飞行日计划表解析结果")
    print(f"  文件: {os.path.basename(filepath)}")
    SEP()

    # codeCipherMap 按读取顺序打印
    print()
    print("【代字代号表】(按读取顺序)")
    SEP("-", 40)
    for i, (dz, dh) in enumerate(parser.code_cipher_map.items(), 1):
        print(f"  {i}. {dz} -> {dh}")
    SEP("-", 40)

    # 时间Map (只打印每个分钟格的第1列，避免重复)
    print()
    print("【时间列映射】(列号 -> 时间)")
    SEP("-", 40)
    prev_time = None
    for col in sorted(parser.time_map.keys()):
        t = parser.time_map[col]
        if t != prev_time:  # 跳过合并单元格第2列的重复
            cl = col_to_letter(col)
            print(f"  col{col:>3} ({cl:>3}列) -> {t}")
            prev_time = t
    SEP("-", 40)

    # 架次信息
    print()
    if not sorties:
        print("  未解析到有效架次。")
        print("  提示: 请检查数据区的代字是否与代字代号表一致。")
    else:
        print(f"【架次信息】共 {len(sorties)} 个架次")
        for i, s in enumerate(sorties, 1):
            sl = col_to_letter(s.start_col)
            el = col_to_letter(s.end_col)
            print()
            SEP("-", 50)
            print(f"  架次 {i}")
            SEP("-", 50)
            print(f"  课目编号 : {s.course}")
            print(f"  起始时间 : {s.start_time}")
            print(f"  结束时间 : {s.end_time}")
            print(f"  所在行   : 第 {s.row} 行")
            print(f"  起始列   : 第 {s.start_col} 列 ({sl}列)")
            print(f"  结束列   : 第 {s.end_col} 列 ({el}列)")
            print(f"  前仓     : {s.front_seat_daizi}{s.front_seat_daihao}"
                  f"  (代字:{s.front_seat_daizi}  代号:{s.front_seat_daihao})")
            if s.is_solo:
                print(f"  后仓     : 无 (单飞)")
            else:
                print(f"  后仓     : {s.rear_seat_daizi}{s.rear_seat_daihao}"
                      f"  (代字:{s.rear_seat_daizi}  代号:{s.rear_seat_daihao})")
            print(f"  飞行类型 : {'单飞 (仅前仓)' if s.is_solo else '带飞 (前仓+后仓)'}")

    print()
    SEP()
    print(f"  解析完成, 共 {len(sorties)} 个架次")
    SEP()


def get_filepath():
    if len(sys.argv) > 1:
        return sys.argv[1].strip().strip('"').strip("'")
    print()
    SEP()
    print("  飞行日计划表解析器")
    print("  请将Excel文件拖入此窗口, 然后按回车")
    SEP()
    print()
    path = input("文件路径: ").strip().strip('"').strip("'")
    return path


def main():
    filepath = get_filepath()

    if not filepath:
        print("错误: 未输入文件路径")
        input("\n按回车键退出...")
        sys.exit(1)

    if not os.path.isfile(filepath):
        print(f"错误: 文件不存在 -> {filepath}")
        input("\n按回车键退出...")
        sys.exit(1)

    try:
        parser = FlightPlanParser(filepath)
        sorties = parser.parse()
        print_results(parser, sorties, filepath)
    except Exception as e:
        print(f"\n解析失败: {e}")
        import traceback
        traceback.print_exc()

    input("\n按回车键退出...")


if __name__ == "__main__":
    main()
