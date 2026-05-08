"""
飞行日计划表解析器 (统一版 - 支持 .xls 和 .xlsx)
====================================================
.xlsx -> openpyxl 直接读取
.xls  -> xlrd 直接读取（含边框颜色）

用法:
    1. 双击运行本脚本，按提示拖入文件
    2. 或在CMD中: python flight_plan_parser.py "C:\路径\计划表.xlsx"
"""

import re
import sys
import os
import json
from dataclasses import dataclass, asdict
from typing import Optional
from collections import OrderedDict


# ============================================================
# 配置常量
# ============================================================
START_ROW = 9
HOUR_TIME_ROW = 7
MIN_TIME_ROW = 8
CODE_CIPHER_GROUPS = [(3, 4), (8, 9), (13, 14)]
COURSE_PATTERN = re.compile(r'^[A-Za-z]{2,4}([-].+)?$')
PERSON_PATTERN = re.compile(r'^([\u4e00-\u9fff])(\d{3,4})$')
DAIZI_PATTERN = re.compile(r'^[\u4e00-\u9fff]$')
DAIHAO_PATTERN = re.compile(r'^\d{3,4}$')
BOLD_STYLES = {'thick', 'medium', 'double'}

XLS_BORDER_STYLE_MAP = {
    0: None, 1: 'thin', 2: 'medium', 3: 'dashed',
    4: 'dotted', 5: 'thick', 6: 'double', 7: 'hair',
}

INDEXED_COLOR_MAP = {
    0: 'black', 1: 'black', 2: 'red', 3: 'green',
    4: 'blue', 5: 'black', 8: 'black', 9: 'black',
    10: 'red', 11: 'green', 12: 'blue', 13: 'black',
    14: 'red', 16: 'red', 18: 'blue', 30: 'red',
    32: 'blue', 36: 'blue', 40: 'blue', 48: 'blue', 53: 'red',
}


# ============================================================
# 颜色识别
# ============================================================
def classify_color_index(idx):
    return INDEXED_COLOR_MAP.get(idx, f'indexed:{idx}')


def classify_color_hex(hex_color):
    if hex_color is None:
        return 'black'
    h = hex_color.upper()
    if h.startswith('INDEXED:'):
        return classify_color_index(int(h.split(':')[1]))
    if h.startswith('THEME:'):
        idx = int(h.split(':')[1])
        theme_map = {0: 'black', 1: 'black', 4: 'blue', 5: 'red', 6: 'green', 9: 'red'}
        return theme_map.get(idx, f'theme:{idx}')
    if len(h) == 8:
        h = h[2:]
    if len(h) != 6:
        return h
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    if b > 150 and r < 100 and g < 100:
        return 'blue'
    if r > 150 and g < 100 and b < 100:
        return 'red'
    if r < 60 and g < 60 and b < 60:
        return 'black'
    return f'rgb({r},{g},{b})'


# ============================================================
# 数据结构
# ============================================================
@dataclass
class BorderRegion:
    row: int
    start_col: int
    end_col: int
    color: str


@dataclass
class FlightSortie:
    course: str
    start_time: str
    end_time: str
    front_seat_daizi: str
    front_seat_daihao: str
    rear_seat_daizi: Optional[str]
    rear_seat_daihao: Optional[str]
    is_solo: bool
    row: int
    start_col: int
    end_col: int
    border_color: str


# ============================================================
# 读取层 - 统一 xls/xlsx 接口
# ============================================================
class XlsxReader:
    def __init__(self, filepath, sheet_name=None):
        import openpyxl
        self.wb = openpyxl.load_workbook(filepath)
        self.ws = self.wb[sheet_name] if sheet_name else self.wb.active
        self.max_row = self.ws.max_row
        self.max_col = self.ws.max_column

    def cell_value(self, row, col):
        return self.ws.cell(row=row, column=col).value

    def bottom_border_info(self, row, col):
        cell = self.ws.cell(row=row, column=col)
        if not cell.border or not cell.border.bottom:
            return None, None
        bottom = cell.border.bottom
        if not bottom.style or bottom.style not in BOLD_STYLES:
            return None, None
        color = bottom.color
        if color is None:
            return bottom.style, 'black'
        if color.type == 'rgb' and color.rgb and color.rgb != '00000000':
            return bottom.style, classify_color_hex(color.rgb)
        if color.type == 'indexed':
            return bottom.style, classify_color_index(color.indexed)
        if color.type == 'theme':
            return bottom.style, classify_color_hex(f'theme:{color.theme}')
        return bottom.style, 'black'

    def close(self):
        self.wb.close()


class XlsReader:
    def __init__(self, filepath, sheet_name=None):
        try:
            import xlrd
        except ImportError:
            print("提示: 读取 .xls 文件需要 xlrd 库，正在安装...")
            import subprocess
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'xlrd', '-q'])
            import xlrd

        self.wb = xlrd.open_workbook(filepath, formatting_info=True)
        if sheet_name:
            self.ws = self.wb.sheet_by_name(sheet_name)
        else:
            self.ws = self.wb.sheet_by_index(0)
        self.max_row = self.ws.nrows
        self.max_col = self.ws.ncols
        self._colour_map = self.wb.colour_map

    def cell_value(self, row, col):
        r, c = row - 1, col - 1
        if r < 0 or r >= self.max_row or c < 0 or c >= self.max_col:
            return None
        v = self.ws.cell_value(r, c)
        if v == '':
            return None
        return v

    def bottom_border_info(self, row, col):
        r, c = row - 1, col - 1
        if r < 0 or r >= self.max_row or c < 0 or c >= self.max_col:
            return None, None
        cell = self.ws.cell(r, c)
        if cell.xf_index is None:
            return None, None
        xf = self.wb.xf_list[cell.xf_index]
        border = xf.border
        line_style = border.bottom_line_style
        style_name = XLS_BORDER_STYLE_MAP.get(line_style)
        if not style_name or style_name not in BOLD_STYLES:
            return None, None
        colour_idx = border.bottom_colour_index
        rgb = self._colour_map.get(colour_idx)
        if rgb and rgb != (0, 0, 0):
            hex_str = f'{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}'
            return style_name, classify_color_hex(hex_str)
        return style_name, classify_color_index(colour_idx)

    def close(self):
        self.wb.release_resources()


def create_reader(filepath, sheet_name=None):
    if filepath.lower().endswith('.xls') and not filepath.lower().endswith('.xlsx'):
        return XlsReader(filepath, sheet_name)
    return XlsxReader(filepath, sheet_name)


# ============================================================
# 解析器
# ============================================================
class FlightPlanParser:
    def __init__(self, filepath: str, sheet_name: str = None):
        self.filepath = filepath
        self.reader = create_reader(filepath, sheet_name)
        self.data_start_col = None
        self.data_end_col = None
        self.time_map = {}
        self.code_cipher_map = OrderedDict()
        self.border_regions = []
        self.metadata = {}
        self.hour_time_row = HOUR_TIME_ROW
        self.min_time_row = MIN_TIME_ROW
        self.start_row = START_ROW

    def cell(self, row, col):
        return self.reader.cell_value(row, col)

    def parse(self) -> list:
        self._find_boundaries()
        self._build_time_map()
        self._build_code_cipher_map()
        self._scan_borders()
        self._extract_metadata()
        return self._extract_sorties()

    def _find_boundaries(self):
        for search_row in [self.hour_time_row] + list(range(5, 12)):
            for col in range(1, self.reader.max_col + 1):
                val = self.cell(search_row, col)
                if val and isinstance(val, str):
                    cleaned = re.sub(r'[\s\u3000]+', '', val)
                    if cleaned in ('飞机', '飞机号码', '飞机号') and self.data_start_col is None:
                        self.data_start_col = col
                        if search_row != self.hour_time_row:
                            self.hour_time_row = search_row
                            self.min_time_row = search_row + 1
                            self.start_row = search_row + 2
                    elif cleaned == '备注' and self.data_end_col is None:
                        self.data_end_col = col
            if self.data_start_col and self.data_end_col:
                break
        if self.data_start_col is None:
            print("\n  [调试] 未找到'飞机'列，打印第5-10行内容:")
            for r in range(5, 11):
                vals = []
                for c in range(1, min(self.reader.max_col + 1, 30)):
                    v = self.cell(r, c)
                    if v is not None:
                        vals.append(f"col{c}={repr(v)}")
                if vals:
                    print(f"    行{r}: {vals}")
            raise ValueError("未找到'飞机'列")
        if self.data_end_col is None:
            raise ValueError("未找到'备注'列")

    def _build_time_map(self):
        current_hour = None
        for col in range(self.data_start_col - 1, 0, -1):
            h = self.cell(self.hour_time_row, col)
            if h and isinstance(h, str) and ':' in h:
                current_hour = int(h.split(':')[0].strip())
                break
        last_base_min = None
        for col in range(self.data_start_col + 1, self.data_end_col):
            h = self.cell(self.hour_time_row, col)
            m = self.cell(self.min_time_row, col)
            if h and isinstance(h, str) and ':' in h:
                current_hour = int(h.split(':')[0].strip())
            if m is not None and current_hour is not None:
                base_min = int(m)
                self.time_map[col] = f"{current_hour}:{str(base_min).zfill(2)}"
                last_base_min = base_min
            elif last_base_min is not None and current_hour is not None:
                next_min = last_base_min + 5
                hr = current_hour
                if next_min >= 60:
                    next_min -= 60
                    hr += 1
                self.time_map[col] = f"{hr}:{str(next_min).zfill(2)}"
                last_base_min = None

    def _build_code_cipher_map(self):
        for row in range(self.start_row, self.reader.max_row + 1):
            for dz_col, dh_col in CODE_CIPHER_GROUPS:
                dz = self.cell(row, dz_col)
                dh = self.cell(row, dh_col)
                if dz is not None and dh is not None:
                    self.code_cipher_map[str(dz).strip()] = str(dh).strip()

    def _scan_borders(self):
        self.border_regions = []
        for row in range(self.start_row, self.reader.max_row + 1):
            bold_cells = []
            for col in range(1, self.reader.max_col + 1):
                style, color = self.reader.bottom_border_info(row, col)
                if style and color in ('blue', 'red'):
                    bold_cells.append((col, color))
            i = 0
            while i < len(bold_cells):
                col, color = bold_cells[i]
                start_col = col
                end_col = col
                j = i + 1
                while j < len(bold_cells):
                    next_col, next_color = bold_cells[j]
                    if next_color == color and next_col == end_col + 1:
                        end_col = next_col
                        j += 1
                    else:
                        break
                if end_col - start_col >= 1:
                    self.border_regions.append(BorderRegion(
                        row=row, start_col=start_col, end_col=end_col, color=color))
                i = j

    def _extract_metadata(self):
        meta = {}
        for row in range(1, 6):
            for col in range(1, self.reader.max_col + 1):
                v = self.cell(row, col)
                if v and isinstance(v, str) and '计划表' in v:
                    meta['标题'] = v.strip()
                    break
            if '标题' in meta:
                break
        for row in range(1, 6):
            for col in range(1, 20):
                v = self.cell(row, col)
                if v and isinstance(v, str) and '批' in v and '准' in v:
                    for c2 in range(col + 1, col + 20):
                        v2 = self.cell(row, c2)
                        if v2 is not None and str(v2).strip():
                            meta['批准'] = str(v2).strip()
                            break
                    break
        for row in range(1, 6):
            for col in range(1, 20):
                v = self.cell(row, col)
                if v and isinstance(v, str) and '年' in v and '月' in v and '日' in v:
                    year_val = None
                    for c2 in range(col - 1, 0, -1):
                        v2 = self.cell(row, c2)
                        if v2 is not None:
                            year_val = str(v2).strip()
                            break
                    date_str = v.strip()
                    if year_val:
                        date_str = year_val + date_str
                    meta['年月日'] = re.sub(r'\s+', '', date_str)
                    break
        label_keys = {
            '塔台呼号': '塔台呼号', '定向台呼号': '定向台呼号',
            '指挥所呼号': '指挥所呼号', '敌我识别信号': '敌我识别信号',
            '天亮时刻': '天亮时刻', '日出时刻': '日出时刻',
            '日没时刻': '日没时刻', '天黑时刻': '天黑时刻',
        }
        for row in range(1, 6):
            for col in range(1, self.reader.max_col + 1):
                v = self.cell(row, col)
                if v and isinstance(v, str):
                    v_cleaned = re.sub(r'[\s\u3000]+', '', v)
                    if v_cleaned in label_keys:
                        key = label_keys[v_cleaned]
                        for c2 in range(col + 1, col + 20):
                            v2 = self.cell(row, c2)
                            if v2 is not None and str(v2).strip():
                                meta[key] = str(v2).strip()
                                break
        if self.data_end_col:
            rv = self.cell(self.start_row, self.data_end_col)
            if rv and str(rv).strip():
                meta['备注'] = str(rv).strip()
        sign_labels = ('飞行指挥员', '塔台飞行管制员', '航空军医')
        for row in range(self.reader.max_row, max(self.reader.max_row - 5, 0), -1):
            for col in range(1, self.reader.max_col + 1):
                v = self.cell(row, col)
                if v and isinstance(v, str):
                    v_clean = re.sub(r'[\s\u3000]+', '', v)
                    if v_clean in sign_labels:
                        if v_clean in meta:
                            continue
                        names = []
                        for c2 in range(col + 1, col + 20):
                            v2 = self.cell(row, c2)
                            if v2 is not None and str(v2).strip():
                                val = str(v2).strip()
                                if re.sub(r'[\s\u3000]+', '', val) in sign_labels:
                                    continue
                                names.append(val)
                        meta[v_clean] = '、'.join(names) if names else None
        self.metadata = meta

    def _extract_sorties(self) -> list:
        results = []
        for region in self.border_regions:
            sortie = self._parse_sortie_from_region(region)
            if sortie and sortie.course != "未识别":
                results.append(sortie)
        return results

    def _parse_sortie_from_region(self, region):
        row, sc, ec = region.row, region.start_col, region.end_col
        course = None
        for sr in (row, row - 1):
            for col in range(sc, ec + 1):
                val = self.cell(sr, col)
                if val is not None:
                    val_str = str(val).strip()
                    if COURSE_PATTERN.match(val_str):
                        course = val_str
                        break
            if course:
                break
        if not course:
            course = "未识别"
        start_time = self._find_nearest_time(sc)
        end_time = self._find_nearest_time(ec + 1)
        front_dz, front_dh, front_col = self._find_person(row, ec + 1)
        rear_dz = rear_dh = None
        is_solo = True
        if front_dz and front_col:
            for offset in (1, 2):
                r_dz, r_dh, _ = self._find_person(row + offset, front_col)
                if r_dz:
                    rear_dz, rear_dh = r_dz, r_dh
                    is_solo = False
                    break
        return FlightSortie(
            course=course, start_time=start_time, end_time=end_time,
            front_seat_daizi=front_dz, front_seat_daihao=front_dh,
            rear_seat_daizi=rear_dz, rear_seat_daihao=rear_dh,
            is_solo=is_solo, row=row, start_col=sc, end_col=ec,
            border_color=region.color)

    def _find_nearest_time(self, col):
        if col in self.time_map:
            return self.time_map[col]
        for c in range(col, 0, -1):
            if c in self.time_map:
                return self.time_map[c]
        return "未知"

    def _find_person(self, row, start_col):
        for col in range(start_col, start_col + 3):
            val = self.cell(row, col)
            if val is None:
                continue
            val_str = str(val).strip()
            match = PERSON_PATTERN.match(val_str)
            if match:
                return match.group(1), match.group(2), col
            if DAIZI_PATTERN.match(val_str):
                nv = self.cell(row, col + 1)
                if nv is not None and DAIHAO_PATTERN.match(str(nv).strip()):
                    return val_str, str(nv).strip(), col
        return None, None, None


# ============================================================
# 输出
# ============================================================
def col_to_letter(n):
    r = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        r = chr(65 + rem) + r
    return r


def SEP(c="=", n=60):
    print(c * n)


def print_results(parser, sorties, filepath):
    print()
    SEP()
    print(f"  飞行日计划表解析结果")
    print(f"  文件: {os.path.basename(filepath)}")
    SEP()
    if parser.metadata:
        print()
        print("【表头/表尾元数据】")
        SEP("-", 50)
        for k, v in parser.metadata.items():
            print(f"  {k}: {v}")
        SEP("-", 50)
    print()
    print("【代字代号表】")
    SEP("-", 40)
    for i, (dz, dh) in enumerate(parser.code_cipher_map.items(), 1):
        print(f"  {i}. {dz} -> {dh}")
    SEP("-", 40)
    print()
    print(f"【边框识别】共 {len(parser.border_regions)} 个区域")
    SEP("-", 50)
    for i, r in enumerate(parser.border_regions, 1):
        sl, el = col_to_letter(r.start_col), col_to_letter(r.end_col)
        print(f"  {i}. 行{r.row}  {sl}{r.row}:{el}{r.row}  列{r.start_col}-{r.end_col}  [{r.color}]")
    SEP("-", 50)
    print()
    if not sorties:
        print("  未解析到有效架次。")
    else:
        print(f"【架次信息】共 {len(sorties)} 个架次")
        for i, s in enumerate(sorties, 1):
            sl, el = col_to_letter(s.start_col), col_to_letter(s.end_col)
            print()
            SEP("-", 50)
            print(f"  架次 {i}")
            SEP("-", 50)
            print(f"  课目编号 : {s.course}")
            print(f"  起始时间 : {s.start_time}")
            print(f"  结束时间 : {s.end_time}")
            print(f"  所在行   : 第 {s.row} 行")
            print(f"  列范围   : {s.start_col}({sl}) - {s.end_col}({el})")
            print(f"  边框颜色 : {s.border_color}")
            if s.front_seat_daizi:
                print(f"  前仓     : {s.front_seat_daizi}{s.front_seat_daihao}")
            else:
                print(f"  前仓     : 未找到")
            if s.is_solo:
                print(f"  后仓     : 无 (单飞)")
            else:
                print(f"  后仓     : {s.rear_seat_daizi}{s.rear_seat_daihao}")
            print(f"  类型     : {'单飞' if s.is_solo else '带飞'}")
    print()
    SEP()
    print(f"  解析完成, 共 {len(sorties)} 个架次")
    SEP()


def build_json(parser, sorties, filepath):
    return {
        "文件": os.path.basename(filepath),
        "元数据": parser.metadata,
        "代字代号表": dict(parser.code_cipher_map),
        "架次总数": len(sorties),
        "架次列表": [asdict(s) for s in sorties],
    }


def get_filepath():
    if len(sys.argv) > 1:
        return sys.argv[1].strip().strip('"').strip("'")
    print()
    SEP()
    print("  飞行日计划表解析器 (支持 .xls / .xlsx)")
    print("  请将Excel文件拖入此窗口, 然后按回车")
    SEP()
    print()
    return input("文件路径: ").strip().strip('"').strip("'")


def main():
    filepath = get_filepath()
    if not filepath:
        print("错误: 未输入文件路径")
        input("\n按回车键退出...")
        sys.exit(1)
    if not os.path.isfile(filepath):
        print(f"错误: 文件不存在 -> {filepath}")
        input("\n按回车键退出...")
        sys.exit(1)
    ext = os.path.splitext(filepath)[1].lower()
    if ext not in ('.xls', '.xlsx'):
        print(f"错误: 不支持的文件格式 '{ext}'")
        input("\n按回车键退出...")
        sys.exit(1)

    try:
        parser = FlightPlanParser(filepath)
        sorties = parser.parse()
        print_results(parser, sorties, filepath)
        data = build_json(parser, sorties, filepath)
        json_path = os.path.splitext(filepath)[0] + '_解析结果.json'
        try:
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            print(f"\n  JSON已保存: {json_path}")
        except OSError:
            json_path = os.path.join('.', os.path.basename(json_path))
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            print(f"\n  JSON已保存: {os.path.abspath(json_path)}")
    except Exception as e:
        print(f"\n解析失败: {e}")
        import traceback
        traceback.print_exc()
    input("\n按回车键退出...")


if __name__ == "__main__":
    main()
