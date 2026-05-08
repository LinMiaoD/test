"""
飞行日计划表解析器 (边框识别版)
================================
通过识别蓝色/红色加粗底边框确定每个架次的起止列范围，
再映射到时间和人员信息。

用法:
    1. 双击运行本脚本，按提示拖入文件
    2. 或在CMD中: python flight_plan_parser.py "C:\路径\计划表.xlsx"
"""

import re
import sys
import os
import json
from dataclasses import dataclass, asdict
from typing import Optional
from collections import OrderedDict

import openpyxl
from openpyxl.styles.colors import Color


# ============================================================
# 配置常量
# ============================================================
START_ROW = 9
HOUR_TIME_ROW = 7
MIN_TIME_ROW = 8
CODE_CIPHER_START_ROW = 9
CODE_CIPHER_GROUPS = [(3, 4), (8, 9), (13, 14)]
COURSE_PATTERN = re.compile(r'^[A-Za-z]{2,4}([-].+)?$')
PERSON_PATTERN = re.compile(r'^([\u4e00-\u9fff])(\d{3,4})$')
DAIZI_PATTERN = re.compile(r'^[\u4e00-\u9fff]$')
DAIHAO_PATTERN = re.compile(r'^\d{3,4}$')

# 加粗边框样式
BOLD_STYLES = {'thick', 'medium', 'double'}


# ============================================================
# 边框颜色识别
# ============================================================
def get_color_hex(color):
    if color is None:
        return None
    if color.type == 'rgb' and color.rgb and color.rgb != '00000000':
        return color.rgb
    if color.type == 'indexed':
        return f'indexed:{color.indexed}'
    if color.type == 'theme':
        return f'theme:{color.theme}'
    return None


def classify_color(hex_color):
    if hex_color is None:
        return 'black'
    h = hex_color.upper()

    if h.startswith('INDEXED:'):
        idx = int(h.split(':')[1])
        indexed_map = {
            0: 'black', 1: 'black',
            2: 'red', 3: 'green',
            4: 'blue', 5: 'black',
            8: 'black', 9: 'black',
            10: 'red', 11: 'green',
            12: 'blue', 13: 'black',
            14: 'red', 16: 'red',
            18: 'blue', 30: 'red',
            32: 'blue', 36: 'blue',
            40: 'blue', 48: 'blue',
            53: 'red',
        }
        return indexed_map.get(idx, f'indexed:{idx}')

    if h.startswith('THEME:'):
        idx = int(h.split(':')[1])
        theme_map = {
            0: 'black', 1: 'black',
            4: 'blue', 5: 'red',
            6: 'green', 9: 'red',
        }
        return theme_map.get(idx, f'theme:{idx}')

    if len(h) == 8:
        h = h[2:]
    if len(h) != 6:
        return h
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    if b > 150 and r < 100 and g < 100:
        return 'blue'
    if r > 150 and g < 100 and b < 100:
        return 'red'
    if r < 60 and g < 60 and b < 60:
        return 'black'
    return f'rgb({r},{g},{b})'


# ============================================================
# 数据结构
# ============================================================
@dataclass
class BorderRegion:
    row: int
    start_col: int
    end_col: int
    color: str


@dataclass
class FlightSortie:
    course: str
    start_time: str
    end_time: str
    front_seat_daizi: str
    front_seat_daihao: str
    rear_seat_daizi: Optional[str]
    rear_seat_daihao: Optional[str]
    is_solo: bool
    row: int
    start_col: int
    end_col: int
    border_color: str


# ============================================================
# xls 边框转换辅助
# ============================================================
def _convert_xls_border(xls_wb, xf):
    """将 xlrd 的 XF 格式转换为 openpyxl Border 对象"""
    from openpyxl.styles.borders import Border, Side

    # xlrd 边框样式映射到 openpyxl
    STYLE_MAP = {
        0: None, 1: 'thin', 2: 'medium', 3: 'dashed',
        4: 'dotted', 5: 'thick', 6: 'double', 7: 'hair',
    }

    def _make_side(line_style, colour_index):
        style = STYLE_MAP.get(line_style)
        if not style:
            return Side()
        # xlrd 颜色索引转为 indexed 颜色
        from openpyxl.styles.colors import Color
        color = Color(indexed=colour_index) if colour_index else None
        return Side(style=style, color=color)

    try:
        border_fmt = xf.border
        top = _make_side(border_fmt.top_line_style, border_fmt.top_colour_index)
        bottom = _make_side(border_fmt.bottom_line_style, border_fmt.bottom_colour_index)
        left = _make_side(border_fmt.left_line_style, border_fmt.left_colour_index)
        right = _make_side(border_fmt.right_line_style, border_fmt.right_colour_index)
        return Border(top=top, bottom=bottom, left=left, right=right)
    except Exception:
        return None


# ============================================================
# 解析器
# ============================================================
class FlightPlanParser:
    def __init__(self, filepath: str, sheet_name: str = None):
        self.filepath = filepath
        self.converted_path = None

        # .xls 文件先转换为 .xlsx
        if filepath.lower().endswith('.xls') and not filepath.lower().endswith('.xlsx'):
            self.converted_path = self._convert_xls_to_xlsx(filepath)
            self.wb = openpyxl.load_workbook(self.converted_path)
        else:
            self.wb = openpyxl.load_workbook(filepath)

        self.ws = self.wb[sheet_name] if sheet_name else self.wb.active
        self.data_start_col = None
        self.data_end_col = None
        self.time_map = {}
        self.code_cipher_map = OrderedDict()
        self.border_regions = []
        self.metadata = {}

    @staticmethod
    def _convert_xls_to_xlsx(xls_path):
        """将 .xls 转换为 .xlsx（需要 xlrd 和 openpyxl）"""
        try:
            import xlrd
        except ImportError:
            print("提示: 读取 .xls 文件需要 xlrd 库，正在安装...")
            import subprocess
            subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'xlrd', '-q'])
            import xlrd

        xlsx_path = xls_path + 'x'  # .xls -> .xlsx
        xls_wb = xlrd.open_workbook(xls_path, formatting_info=True)
        xlsx_wb = openpyxl.Workbook()

        for i, sheet_name in enumerate(xls_wb.sheet_names()):
            xls_sheet = xls_wb.sheet_by_name(sheet_name)
            if i == 0:
                xlsx_sheet = xlsx_wb.active
                xlsx_sheet.title = sheet_name
            else:
                xlsx_sheet = xlsx_wb.create_sheet(sheet_name)

            for row in range(xls_sheet.nrows):
                for col in range(xls_sheet.ncols):
                    cell = xls_sheet.cell(row, col)
                    xlsx_cell = xlsx_sheet.cell(row=row + 1, column=col + 1, value=cell.value)

                    # 转换边框样式
                    xf_index = cell.xf_index
                    if xf_index is not None:
                        xf = xls_wb.xf_list[xf_index]
                        border = _convert_xls_border(xls_wb, xf)
                        if border:
                            xlsx_cell.border = border

        xlsx_wb.save(xlsx_path)
        xlsx_wb.close()
        xls_wb.release_resources()
        print(f"  已将 .xls 转换为 .xlsx: {xlsx_path}")
        return xlsx_path

    def parse(self) -> list:
        self._find_boundaries()
        self._build_time_map()
        self._build_code_cipher_map()
        self._scan_borders()
        self._extract_metadata()
        return self._extract_sorties()

    # ---------- 边界检测 ----------
    def _find_boundaries(self):
        # 先在 HOUR_TIME_ROW 找，找不到就扩大到前后几行
        for search_row in [HOUR_TIME_ROW] + list(range(5, 12)):
            for col in range(1, self.ws.max_column + 1):
                val = self.ws.cell(row=search_row, column=col).value
                if val and isinstance(val, str):
                    cleaned = re.sub(r'\s+', '', val)
                    if cleaned in ('飞机', '飞机号码', '飞机号') and self.data_start_col is None:
                        self.data_start_col = col
                        # 如果不在默认行，更新时间行号
                        if search_row != HOUR_TIME_ROW:
                            global HOUR_TIME_ROW, MIN_TIME_ROW, START_ROW
                            MIN_TIME_ROW = search_row + 1
                            HOUR_TIME_ROW = search_row
                            START_ROW = search_row + 2
                    elif cleaned == '备注' and self.data_end_col is None:
                        self.data_end_col = col
            if self.data_start_col and self.data_end_col:
                break

        if self.data_start_col is None:
            # 调试: 打印第5-10行所有有值单元格帮助定位
            print("\n  [调试] 未找到'飞机'列，打印第5-10行内容:")
            for r in range(5, 11):
                vals = []
                for c in range(1, min(self.ws.max_column + 1, 30)):
                    v = self.ws.cell(row=r, column=c).value
                    if v is not None:
                        vals.append(f"col{c}={repr(v)}")
                if vals:
                    print(f"    行{r}: {vals}")
            raise ValueError("未找到'飞机'列，请检查上方调试信息")
        if self.data_end_col is None:
            raise ValueError("未找到'备注'列")

    # ---------- 时间映射 ----------
    def _build_time_map(self):
        current_hour = None
        for col in range(self.data_start_col - 1, 0, -1):
            h = self.ws.cell(row=HOUR_TIME_ROW, column=col).value
            if h and isinstance(h, str) and ':' in h:
                current_hour = int(h.split(':')[0].strip())
                break
        last_base_min = None
        for col in range(self.data_start_col + 1, self.data_end_col):
            h = self.ws.cell(row=HOUR_TIME_ROW, column=col).value
            m = self.ws.cell(row=MIN_TIME_ROW, column=col).value
            if h and isinstance(h, str) and ':' in h:
                current_hour = int(h.split(':')[0].strip())
            if m is not None and current_hour is not None:
                base_min = int(m)
                self.time_map[col] = f"{current_hour}:{str(base_min).zfill(2)}"
                last_base_min = base_min
            elif last_base_min is not None and current_hour is not None:
                next_min = last_base_min + 5
                hr = current_hour
                if next_min >= 60:
                    next_min -= 60
                    hr += 1
                self.time_map[col] = f"{hr}:{str(next_min).zfill(2)}"
                last_base_min = None

    # ---------- 代字代号表 ----------
    def _build_code_cipher_map(self):
        for row in range(CODE_CIPHER_START_ROW, self.ws.max_row + 1):
            for dz_col, dh_col in CODE_CIPHER_GROUPS:
                dz = self.ws.cell(row=row, column=dz_col).value
                dh = self.ws.cell(row=row, column=dh_col).value
                if dz is not None and dh is not None:
                    self.code_cipher_map[str(dz).strip()] = str(dh).strip()

    # ---------- 提取表头/表尾元数据 ----------
    def _extract_metadata(self):
        ws = self.ws
        meta = {}

        # --- 表头标题: 扫描前5行找 "XX计划表" 样式的标题 ---
        for row in range(1, 6):
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=row, column=col).value
                if v and isinstance(v, str) and '计划表' in v:
                    meta['标题'] = v.strip()
                    break
            if '标题' in meta:
                break

        # --- 批准 ---
        for row in range(1, 6):
            for col in range(1, 20):
                v = ws.cell(row=row, column=col).value
                if v and isinstance(v, str) and '批' in v and '准' in v:
                    # 批准人在后面几列
                    for c2 in range(col + 1, col + 10):
                        v2 = ws.cell(row=row, column=c2).value
                        if v2 is not None and str(v2).strip():
                            meta['批准'] = str(v2).strip()
                            break
                    break

        # --- 年月日 ---
        for row in range(1, 6):
            for col in range(1, 20):
                v = ws.cell(row=row, column=col).value
                if v and isinstance(v, str) and '年' in v and '月' in v and '日' in v:
                    # 年份可能在前面的列
                    year_val = None
                    for c2 in range(col - 1, 0, -1):
                        v2 = ws.cell(row=row, column=c2).value
                        if v2 is not None:
                            year_val = str(v2).strip()
                            break
                    date_str = v.strip()
                    if year_val:
                        date_str = year_val + date_str
                    # 清理多余空格
                    date_str = re.sub(r'\s+', '', date_str)
                    meta['年月日'] = date_str
                    break

        # --- 右上角信息: 塔台呼号、定向台呼号、指挥所呼号、敌我识别信号 ---
        # --- 以及: 天亮时刻、日出时刻、日没时刻、天黑时刻 ---
        label_keys = {
            '塔台呼号': '塔台呼号', '定向台呼号': '定向台呼号',
            '指挥所呼号': '指挥所呼号', '敌我识别信号': '敌我识别信号',
            '天亮时刻': '天亮时刻', '日出时刻': '日出时刻',
            '日没时刻': '日没时刻', '天黑时刻': '天黑时刻',
        }
        for row in range(1, 6):
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=row, column=col).value
                if v and isinstance(v, str):
                    v_clean = v.strip()
                    if v_clean in label_keys:
                        key = label_keys[v_clean]
                        # 值在后面几列
                        for c2 in range(col + 1, col + 10):
                            v2 = ws.cell(row=row, column=c2).value
                            if v2 is not None and str(v2).strip():
                                meta[key] = str(v2).strip()
                                break

        # --- 备注 ---
        if self.data_end_col:
            remarks_val = ws.cell(row=START_ROW, column=self.data_end_col).value
            if remarks_val and str(remarks_val).strip():
                meta['备注'] = str(remarks_val).strip()

        # --- 底部签名区: 飞行指挥员、塔台飞行管制员、航空军医 ---
        sign_labels = ('飞行指挥员', '塔台飞行管制员', '航空军医')
        for row in range(ws.max_row, max(ws.max_row - 5, 0), -1):
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=row, column=col).value
                if v and isinstance(v, str):
                    v_clean = v.strip()
                    if v_clean in sign_labels:
                        if v_clean in meta:
                            continue
                        names = []
                        for c2 in range(col + 1, col + 10):
                            v2 = ws.cell(row=row, column=c2).value
                            if v2 is not None and str(v2).strip():
                                val = str(v2).strip()
                                # 跳过和标签相同的值（合并单元格拆开导致的重复）
                                if val in sign_labels:
                                    continue
                                names.append(val)
                        meta[v_clean] = '、'.join(names) if names else None

        self.metadata = meta

    # ---------- 边框扫描 ----------
    def _scan_borders(self):
        self.border_regions = []
        for row_cells in self.ws.iter_rows(min_row=START_ROW):
            row_num = row_cells[0].row
            # 收集本行加粗底边框单元格
            bold_cells = []
            for cell in row_cells:
                if cell.border and cell.border.bottom and cell.border.bottom.style in BOLD_STYLES:
                    hex_c = get_color_hex(cell.border.bottom.color)
                    color = classify_color(hex_c)
                    if color in ('blue', 'red'):
                        bold_cells.append((cell.column, color))

            # 找连续段（同色），至少2列
            i = 0
            while i < len(bold_cells):
                col, color = bold_cells[i]
                start_col = col
                end_col = col
                j = i + 1
                while j < len(bold_cells):
                    next_col, next_color = bold_cells[j]
                    if next_color == color and next_col == end_col + 1:
                        end_col = next_col
                        j += 1
                    else:
                        break
                if end_col - start_col >= 1:
                    self.border_regions.append(BorderRegion(
                        row=row_num, start_col=start_col,
                        end_col=end_col, color=color,
                    ))
                i = j

    # ---------- 提取架次 ----------
    def _extract_sorties(self) -> list:
        results = []
        for region in self.border_regions:
            sortie = self._parse_sortie_from_region(region)
            if sortie and sortie.course != "未识别":
                results.append(sortie)
        return results

    def _parse_sortie_from_region(self, region: BorderRegion):
        row = region.row
        sc = region.start_col
        ec = region.end_col

        # 1. 课目: 先在当前行(row)找，找不到再在上一行(row-1)找
        course = None
        for search_row in (row, row - 1):
            for col in range(sc, ec + 1):
                val = self.ws.cell(row=search_row, column=col).value
                if val is not None:
                    val_str = str(val).strip()
                    if COURSE_PATTERN.match(val_str):
                        course = val_str
                        break
            if course:
                break
        if course is None:
            course = "未识别"

        # 2. 时间: start_col -> 开始时间, end_col+1 -> 结束时间(下一列开始即本段结束)
        start_time = self._find_nearest_time(sc)
        end_time = self._find_nearest_time(ec + 1)

        # 3. 人员: 在 end_col+1 开始找前仓（ec+1没有就ec+2）
        front_dz, front_dh, front_col = self._find_person(row, ec + 1)

        # 4. 后仓: 从同一列开始找，row+1 或 row+2
        rear_dz = rear_dh = None
        is_solo = True
        if front_dz and front_col:
            for offset in (1, 2):
                r_dz, r_dh, _ = self._find_person(row + offset, front_col)
                if r_dz:
                    rear_dz, rear_dh = r_dz, r_dh
                    is_solo = False
                    break

        return FlightSortie(
            course=course, start_time=start_time, end_time=end_time,
            front_seat_daizi=front_dz, front_seat_daihao=front_dh,
            rear_seat_daizi=rear_dz, rear_seat_daihao=rear_dh,
            is_solo=is_solo, row=row, start_col=sc, end_col=ec,
            border_color=region.color,
        )

    def _find_nearest_time(self, col):
        """找col或最近的有时间映射的列"""
        if col in self.time_map:
            return self.time_map[col]
        # 往前找最近的
        for c in range(col, 0, -1):
            if c in self.time_map:
                return self.time_map[c]
        return "未知"

    def _find_person(self, row, start_col):
        """
        从 start_col 开始，在最多3列范围内找代字+代号。
        返回 (代字, 代号, 代字所在列) 或 (None, None, None)
        情况1: 同一列 "X1234" 格式
        情况2: 代字在一列(汉字)，代号在代字列+1
        """
        for col in range(start_col, start_col + 3):
            val = self.ws.cell(row=row, column=col).value
            if val is None:
                continue
            val_str = str(val).strip()

            # 情况1: 代字代号在同一单元格 "X1234"
            match = PERSON_PATTERN.match(val_str)
            if match:
                dz, dh = match.group(1), match.group(2)
                return dz, dh, col

            # 情况2: 当前单元格只有代字(单个汉字)，代字列+1是代号(数字)
            if DAIZI_PATTERN.match(val_str):
                next_val = self.ws.cell(row=row, column=col + 1).value
                if next_val is not None:
                    next_str = str(next_val).strip()
                    if DAIHAO_PATTERN.match(next_str):
                        return val_str, next_str, col

        return None, None, None


# ============================================================
# 输出
# ============================================================
def col_to_letter(col_num):
    result = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result


def SEP(char="=", length=60):
    print(char * length)


def print_results(parser, sorties, filepath):
    print()
    SEP()
    print(f"  飞行日计划表解析结果 (边框识别版)")
    print(f"  文件: {os.path.basename(filepath)}")
    SEP()

    # 表头元数据
    if parser.metadata:
        print()
        print("【表头/表尾元数据】")
        SEP("-", 50)
        for k, v in parser.metadata.items():
            print(f"  {k}: {v}")
        SEP("-", 50)

    # 代字代号表
    print()
    print("【代字代号表】(按读取顺序)")
    SEP("-", 40)
    for i, (dz, dh) in enumerate(parser.code_cipher_map.items(), 1):
        print(f"  {i}. {dz} -> {dh}")
    SEP("-", 40)

    # 时间映射
    print()
    print("【时间列映射】(列号 -> 时间)")
    SEP("-", 40)
    prev_time = None
    for col in sorted(parser.time_map.keys()):
        t = parser.time_map[col]
        if t != prev_time:
            cl = col_to_letter(col)
            print(f"  col{col:>3} ({cl:>3}列) -> {t}")
            prev_time = t
    SEP("-", 40)

    # 边框识别结果
    print()
    print(f"【边框识别】共 {len(parser.border_regions)} 个区域")
    SEP("-", 50)
    for i, r in enumerate(parser.border_regions, 1):
        sl = col_to_letter(r.start_col)
        el = col_to_letter(r.end_col)
        print(f"  {i}. 行{r.row}  {sl}{r.row}:{el}{r.row}"
              f"  列{r.start_col}-{r.end_col}  [{r.color}]")
    SEP("-", 50)

    # 架次信息
    print()
    if not sorties:
        print("  未解析到有效架次。")
    else:
        print(f"【架次信息】共 {len(sorties)} 个架次")
        for i, s in enumerate(sorties, 1):
            sl = col_to_letter(s.start_col)
            el = col_to_letter(s.end_col)
            print()
            SEP("-", 50)
            print(f"  架次 {i}")
            SEP("-", 50)
            print(f"  课目编号 : {s.course}")
            print(f"  起始时间 : {s.start_time}")
            print(f"  结束时间 : {s.end_time}")
            print(f"  所在行   : 第 {s.row} 行")
            print(f"  起始列   : 第 {s.start_col} 列 ({sl}列)")
            print(f"  结束列   : 第 {s.end_col} 列 ({el}列)")
            print(f"  边框颜色 : {s.border_color}")
            if s.front_seat_daizi:
                print(f"  前仓     : {s.front_seat_daizi}{s.front_seat_daihao}"
                      f"  (代字:{s.front_seat_daizi}  代号:{s.front_seat_daihao})")
            else:
                print(f"  前仓     : 未找到")
            if s.is_solo:
                print(f"  后仓     : 无 (单飞)")
            else:
                print(f"  后仓     : {s.rear_seat_daizi}{s.rear_seat_daihao}"
                      f"  (代字:{s.rear_seat_daizi}  代号:{s.rear_seat_daihao})")
            print(f"  飞行类型 : {'单飞 (仅前仓)' if s.is_solo else '带飞 (前仓+后仓)'}")

    print()
    SEP()
    print(f"  解析完成, 共 {len(sorties)} 个架次")
    SEP()


def build_json(parser, sorties, filepath):
    """将所有解析结果构建为一个完整的JSON"""
    result = {
        "文件": os.path.basename(filepath),
        "元数据": parser.metadata,
        "代字代号表": dict(parser.code_cipher_map),
        "架次总数": len(sorties),
        "架次列表": [asdict(s) for s in sorties],
    }
    return result


def get_filepath():
    if len(sys.argv) > 1:
        return sys.argv[1].strip().strip('"').strip("'")
    print()
    SEP()
    print("  飞行日计划表解析器 (边框识别版)")
    print("  请将Excel文件拖入此窗口, 然后按回车")
    SEP()
    print()
    path = input("文件路径: ").strip().strip('"').strip("'")
    return path


def main():
    filepath = get_filepath()

    if not filepath:
        print("错误: 未输入文件路径")
        input("\n按回车键退出...")
        sys.exit(1)

    if not os.path.isfile(filepath):
        print(f"错误: 文件不存在 -> {filepath}")
        input("\n按回车键退出...")
        sys.exit(1)

    try:
        parser = FlightPlanParser(filepath)
        sorties = parser.parse()
        print_results(parser, sorties, filepath)

        # 输出JSON文件
        data = build_json(parser, sorties, filepath)
        json_path = os.path.splitext(filepath)[0] + '_解析结果.json'
        try:
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            print(f"\n  JSON已保存: {json_path}")
        except OSError:
            # 源文件目录不可写时，保存到当前目录
            json_path = os.path.join('.', os.path.basename(json_path))
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            print(f"\n  JSON已保存: {os.path.abspath(json_path)}")
    except Exception as e:
        print(f"\n解析失败: {e}")
        import traceback
        traceback.print_exc()

    input("\n按回车键退出...")


if __name__ == "__main__":
    main()
