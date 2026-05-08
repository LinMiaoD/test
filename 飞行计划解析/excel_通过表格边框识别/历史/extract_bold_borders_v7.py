"""
Excel 加粗边框区域提取工具（逐行识别版）
规则:
  - 逐行扫描，不跨行
  - 以单元格的 底边框(bottom) 是否加粗为判断依据
  - 同一行内连续的加粗底边框单元格为一个区域
  - 底边框不加粗即视为该区域结束
  - 输出: 行号、开始列、结束列、颜色、样式

用法:
  python extract_bold_borders.py <excel文件路径>
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json, sys

BOLD_STYLES = {'thick', 'medium', 'double'}


def get_color_hex(color):
    if color is None:
        return None
    if color.type == 'rgb' and color.rgb and color.rgb != '00000000':
        return color.rgb
    if color.type == 'indexed':
        return f'indexed:{color.indexed}'
    if color.type == 'theme':
        return f'theme:{color.theme}'
    return None


def classify_color(hex_color):
    if hex_color is None:
        return 'black'
    h = hex_color.upper()

    # Excel indexed 调色板映射
    if h.startswith('INDEXED:'):
        idx = int(h.split(':')[1])
        indexed_map = {
            0: 'black',   1: 'black',    # 000000
            2: 'red',     3: 'green',    # FF0000, 00FF00
            4: 'blue',    5: 'black',    # 0000FF, FFFF00
            8: 'black',   9: 'black',    # 000000
            10: 'red',    11: 'green',   # FF0000, 00FF00
            12: 'blue',   13: 'black',   # 0000FF, FFFF00
            14: 'red',    # dark red
            16: 'red',    # dark red
            18: 'blue',   # dark blue
            30: 'red',    # coral/red
            32: 'blue',   # light blue
            36: 'blue',   # sky blue
            40: 'blue',   # medium blue
            48: 'blue',   # dark teal/blue
            53: 'red',    # dark red variant
        }
        return indexed_map.get(idx, f'indexed:{idx}')

    # Excel theme 颜色映射（常见的）
    if h.startswith('THEME:'):
        idx = int(h.split(':')[1])
        theme_map = {
            0: 'black',   1: 'black',
            4: 'blue',    5: 'red',
            6: 'green',   9: 'red',
        }
        return theme_map.get(idx, f'theme:{idx}')

    if len(h) == 8:
        h = h[2:]
    if len(h) != 6:
        return h
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    if b > 150 and r < 100 and g < 100:
        return 'blue'
    if r > 150 and g < 100 and b < 100:
        return 'red'
    if r < 60 and g < 60 and b < 60:
        return 'black'
    if r < 100 and g > 150 and b < 100:
        return 'green'
    return f'rgb({r},{g},{b})'


def scan_bold_borders(filepath, sheet_name=None):
    wb = load_workbook(filepath, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]
        if sheet_name:
            print(f"警告: Sheet '{sheet_name}' 不存在，使用: '{wb.sheetnames[0]}'")

    print(f"Sheet: {ws.title}")
    print(f"数据范围: {ws.dimensions}\n")

    regions = []

    for row in ws.iter_rows():
        # 收集本行中底边框加粗的单元格
        bold_cells = []
        for cell in row:
            if cell.border is None:
                bold_cells.append(None)
                continue
            bottom = cell.border.bottom
            if bottom and bottom.style in BOLD_STYLES:
                hex_c = get_color_hex(bottom.color)
                bold_cells.append({
                    'row': cell.row,
                    'col': cell.column,
                    'style': bottom.style,
                    'color': classify_color(hex_c),
                    'hex': hex_c
                })
            else:
                bold_cells.append(None)

        # 找连续的加粗段
        i = 0
        while i < len(bold_cells):
            if bold_cells[i] is None:
                i += 1
                continue
            # 找到一个加粗单元格，开始一段
            start = bold_cells[i]
            color = start['color']
            style = start['style']
            end = start
            j = i + 1
            # 往后找连续的、同颜色同样式的
            while j < len(bold_cells):
                if (bold_cells[j] is not None
                        and bold_cells[j]['color'] == color
                        and bold_cells[j]['style'] == style):
                    end = bold_cells[j]
                    j += 1
                else:
                    break

            sc = start['col']
            ec = end['col']
            # 只保留跨2列及以上的区域
            if ec - sc < 1:
                i = j
                continue
            # 只统计蓝色和红色，黑色及其他不计入
            if color not in ('blue', 'red'):
                i = j
                continue
            regions.append({
                'row': start['row'],
                'start_col': sc,
                'start_col_letter': get_column_letter(sc),
                'end_col': ec,
                'end_col_letter': get_column_letter(ec),
                'range': f"{get_column_letter(sc)}{start['row']}:{get_column_letter(ec)}{start['row']}",
                'color': color,
                'style': style,
            })
            i = j

    wb.close()
    return regions


def print_results(regions):
    if not regions:
        print("未发现加粗底边框区域")
        return
    print(f"共发现 {len(regions)} 个加粗底边框区域:\n")
    print(f"{'行':<6} {'开始列':<10} {'结束列':<10} {'范围':<14} {'颜色':<12} {'样式'}")
    print('-' * 65)
    for r in regions:
        print(f"{r['row']:<6} "
              f"{r['start_col']}({r['start_col_letter']}){'':^4} "
              f"{r['end_col']}({r['end_col_letter']}){'':^4} "
              f"{r['range']:<14} "
              f"{r['color']:<12} "
              f"{r['style']}")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("用法: python extract_bold_borders.py <excel文件>")
        sys.exit(1)

    filepath = ' '.join(sys.argv[1:])
    results = scan_bold_borders(filepath)
    print_results(results)
    print("\n--- JSON ---")
    print(json.dumps(results, ensure_ascii=False, indent=2))
