"""
飞行日计划表解析器 (边框识别版)
================================
通过识别蓝色/红色加粗底边框确定每个架次的起止列范围，
再映射到时间和人员信息。

用法:
    1. 双击运行本脚本，按提示拖入文件
    2. 或在CMD中: python flight_plan_parser.py "C:\路径\计划表.xlsx"
"""

import re
import sys
import os
from dataclasses import dataclass
from typing import Optional
from collections import OrderedDict

import openpyxl
from openpyxl.styles.colors import Color


# ============================================================
# 配置常量
# ============================================================
START_ROW = 9
HOUR_TIME_ROW = 7
MIN_TIME_ROW = 8
CODE_CIPHER_START_ROW = 9
CODE_CIPHER_GROUPS = [(3, 4), (8, 9), (13, 14)]
COURSE_PATTERN = re.compile(r'^[A-Za-z]{2,4}[-].+$')
PERSON_PATTERN = re.compile(r'^([\u4e00-\u9fff])(\d{3,4})$')
DAIZI_PATTERN = re.compile(r'^[\u4e00-\u9fff]$')
DAIHAO_PATTERN = re.compile(r'^\d{3,4}$')

# 加粗边框样式
BOLD_STYLES = {'thick', 'medium', 'double'}


# ============================================================
# 边框颜色识别
# ============================================================
def get_color_hex(color):
    if color is None:
        return None
    if color.type == 'rgb' and color.rgb and color.rgb != '00000000':
        return color.rgb
    if color.type == 'indexed':
        return f'indexed:{color.indexed}'
    if color.type == 'theme':
        return f'theme:{color.theme}'
    return None


def classify_color(hex_color):
    if hex_color is None:
        return 'black'
    h = hex_color.upper()

    if h.startswith('INDEXED:'):
        idx = int(h.split(':')[1])
        indexed_map = {
            0: 'black', 1: 'black',
            2: 'red', 3: 'green',
            4: 'blue', 5: 'black',
            8: 'black', 9: 'black',
            10: 'red', 11: 'green',
            12: 'blue', 13: 'black',
            14: 'red', 16: 'red',
            18: 'blue', 30: 'red',
            32: 'blue', 36: 'blue',
            40: 'blue', 48: 'blue',
            53: 'red',
        }
        return indexed_map.get(idx, f'indexed:{idx}')

    if h.startswith('THEME:'):
        idx = int(h.split(':')[1])
        theme_map = {
            0: 'black', 1: 'black',
            4: 'blue', 5: 'red',
            6: 'green', 9: 'red',
        }
        return theme_map.get(idx, f'theme:{idx}')

    if len(h) == 8:
        h = h[2:]
    if len(h) != 6:
        return h
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    if b > 150 and r < 100 and g < 100:
        return 'blue'
    if r > 150 and g < 100 and b < 100:
        return 'red'
    if r < 60 and g < 60 and b < 60:
        return 'black'
    return f'rgb({r},{g},{b})'


# ============================================================
# 数据结构
# ============================================================
@dataclass
class BorderRegion:
    row: int
    start_col: int
    end_col: int
    color: str


@dataclass
class FlightSortie:
    course: str
    start_time: str
    end_time: str
    front_seat_daizi: str
    front_seat_daihao: str
    rear_seat_daizi: Optional[str]
    rear_seat_daihao: Optional[str]
    is_solo: bool
    row: int
    start_col: int
    end_col: int
    border_color: str


# ============================================================
# 解析器
# ============================================================
class FlightPlanParser:
    def __init__(self, filepath: str, sheet_name: str = None):
        self.filepath = filepath
        self.wb = openpyxl.load_workbook(filepath)
        self.ws = self.wb[sheet_name] if sheet_name else self.wb.active
        self.data_start_col = None
        self.data_end_col = None
        self.time_map = {}
        self.code_cipher_map = OrderedDict()
        self.border_regions = []

    def parse(self) -> list:
        self._find_boundaries()
        self._build_time_map()
        self._build_code_cipher_map()
        self._scan_borders()
        return self._extract_sorties()

    # ---------- 边界检测 ----------
    def _find_boundaries(self):
        for col in range(1, self.ws.max_column + 1):
            val = self.ws.cell(row=HOUR_TIME_ROW, column=col).value
            if val == '飞机':
                self.data_start_col = col
            elif val == '备注':
                self.data_end_col = col
        if self.data_start_col is None:
            raise ValueError("未找到'飞机'列")
        if self.data_end_col is None:
            raise ValueError("未找到'备注'列")

    # ---------- 时间映射 ----------
    def _build_time_map(self):
        current_hour = None
        for col in range(self.data_start_col - 1, 0, -1):
            h = self.ws.cell(row=HOUR_TIME_ROW, column=col).value
            if h and isinstance(h, str) and ':' in h:
                current_hour = int(h.split(':')[0].strip())
                break
        last_base_min = None
        for col in range(self.data_start_col + 1, self.data_end_col):
            h = self.ws.cell(row=HOUR_TIME_ROW, column=col).value
            m = self.ws.cell(row=MIN_TIME_ROW, column=col).value
            if h and isinstance(h, str) and ':' in h:
                current_hour = int(h.split(':')[0].strip())
            if m is not None and current_hour is not None:
                base_min = int(m)
                self.time_map[col] = f"{current_hour}:{str(base_min).zfill(2)}"
                last_base_min = base_min
            elif last_base_min is not None and current_hour is not None:
                next_min = last_base_min + 5
                hr = current_hour
                if next_min >= 60:
                    next_min -= 60
                    hr += 1
                self.time_map[col] = f"{hr}:{str(next_min).zfill(2)}"
                last_base_min = None

    # ---------- 代字代号表 ----------
    def _build_code_cipher_map(self):
        for row in range(CODE_CIPHER_START_ROW, self.ws.max_row + 1):
            for dz_col, dh_col in CODE_CIPHER_GROUPS:
                dz = self.ws.cell(row=row, column=dz_col).value
                dh = self.ws.cell(row=row, column=dh_col).value
                if dz is not None and dh is not None:
                    self.code_cipher_map[str(dz).strip()] = str(dh).strip()

    # ---------- 边框扫描 ----------
    def _scan_borders(self):
        self.border_regions = []
        for row_cells in self.ws.iter_rows(min_row=START_ROW):
            row_num = row_cells[0].row
            # 收集本行加粗底边框单元格
            bold_cells = []
            for cell in row_cells:
                if cell.border and cell.border.bottom and cell.border.bottom.style in BOLD_STYLES:
                    hex_c = get_color_hex(cell.border.bottom.color)
                    color = classify_color(hex_c)
                    if color in ('blue', 'red'):
                        bold_cells.append((cell.column, color))

            # 找连续段（同色），至少2列
            i = 0
            while i < len(bold_cells):
                col, color = bold_cells[i]
                start_col = col
                end_col = col
                j = i + 1
                while j < len(bold_cells):
                    next_col, next_color = bold_cells[j]
                    if next_color == color and next_col == end_col + 1:
                        end_col = next_col
                        j += 1
                    else:
                        break
                if end_col - start_col >= 1:
                    self.border_regions.append(BorderRegion(
                        row=row_num, start_col=start_col,
                        end_col=end_col, color=color,
                    ))
                i = j

    # ---------- 提取架次 ----------
    def _extract_sorties(self) -> list:
        results = []
        for region in self.border_regions:
            sortie = self._parse_sortie_from_region(region)
            if sortie:
                results.append(sortie)
        return results

    def _parse_sortie_from_region(self, region: BorderRegion):
        row = region.row
        sc = region.start_col
        ec = region.end_col

        # 1. 课目: 边框行的上一行(row-1)，从start_col开始找
        course = None
        for col in range(sc, ec + 1):
            val = self.ws.cell(row=row - 1, column=col).value
            if val is not None:
                val_str = str(val).strip()
                if COURSE_PATTERN.match(val_str):
                    course = val_str
                    break
        if course is None:
            # 调试: 打印 row-1 行从 sc 开始的前几个有值单元格
            debug_vals = []
            for col in range(sc, min(sc + 10, ec + 2)):
                v = self.ws.cell(row=row - 1, column=col).value
                if v is not None:
                    debug_vals.append(f"col{col}={repr(v)}")
            print(f"  [调试] 行{row} 边框({sc}-{ec}) 课目未识别, "
                  f"row-1={row-1} 的值: {debug_vals if debug_vals else '全空'}")
            course = "未识别"

        # 2. 时间: start_col -> 开始时间, end_col+1 -> 结束时间(下一列开始即本段结束)
        start_time = self._find_nearest_time(sc)
        end_time = self._find_nearest_time(ec + 1)

        # 3. 人员: 在 end_col 之后找前仓
        front_dz, front_dh = self._find_person(row, ec + 1)

        # 4. 后仓: 同列，row+1 或 row+2
        rear_dz = rear_dh = None
        is_solo = True
        if front_dz:
            # 确定前仓所在的列（代字列）
            front_person_col = self._find_person_col(row, ec + 1)
            if front_person_col:
                for offset in (1, 2):
                    r_dz, r_dh = self._find_person(row + offset, front_person_col)
                    if r_dz:
                        rear_dz, rear_dh = r_dz, r_dh
                        is_solo = False
                        break

        return FlightSortie(
            course=course, start_time=start_time, end_time=end_time,
            front_seat_daizi=front_dz, front_seat_daihao=front_dh,
            rear_seat_daizi=rear_dz, rear_seat_daihao=rear_dh,
            is_solo=is_solo, row=row, start_col=sc, end_col=ec,
            border_color=region.color,
        )

    def _find_nearest_time(self, col):
        """找col或最近的有时间映射的列"""
        if col in self.time_map:
            return self.time_map[col]
        # 往前找最近的
        for c in range(col, 0, -1):
            if c in self.time_map:
                return self.time_map[c]
        return "未知"

    def _find_person(self, row, start_col):
        """
        从 start_col 开始，在最多3列范围内找代字+代号。
        情况1: 同一列 "X1234" 格式
        情况2: 代字在一列(汉字)，代号在下一列(数字)
        """
        for col in range(start_col, start_col + 3):
            val = self.ws.cell(row=row, column=col).value
            if val is None:
                continue
            val_str = str(val).strip()

            # 情况1: 代字代号在同一单元格 "X1234"
            match = PERSON_PATTERN.match(val_str)
            if match:
                dz, dh = match.group(1), match.group(2)
                if dz in self.code_cipher_map and self.code_cipher_map[dz] == dh:
                    return dz, dh

            # 情况2: 当前单元格只有代字(单个汉字)，下一列是代号(数字)
            if DAIZI_PATTERN.match(val_str):
                next_val = self.ws.cell(row=row, column=col + 1).value
                if next_val is not None:
                    next_str = str(next_val).strip()
                    if DAIHAO_PATTERN.match(next_str):
                        if val_str in self.code_cipher_map and self.code_cipher_map[val_str] == next_str:
                            return val_str, next_str

        return None, None

    def _find_person_col(self, row, start_col):
        """找到人员代字所在的列号，用于定位后仓"""
        for col in range(start_col, start_col + 3):
            val = self.ws.cell(row=row, column=col).value
            if val is None:
                continue
            val_str = str(val).strip()
            if PERSON_PATTERN.match(val_str):
                return col
            if DAIZI_PATTERN.match(val_str):
                return col
        return None


# ============================================================
# 输出
# ============================================================
def col_to_letter(col_num):
    result = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result


def SEP(char="=", length=60):
    print(char * length)


def print_results(parser, sorties, filepath):
    print()
    SEP()
    print(f"  飞行日计划表解析结果 (边框识别版)")
    print(f"  文件: {os.path.basename(filepath)}")
    SEP()

    # 代字代号表
    print()
    print("【代字代号表】(按读取顺序)")
    SEP("-", 40)
    for i, (dz, dh) in enumerate(parser.code_cipher_map.items(), 1):
        print(f"  {i}. {dz} -> {dh}")
    SEP("-", 40)

    # 时间映射
    print()
    print("【时间列映射】(列号 -> 时间)")
    SEP("-", 40)
    prev_time = None
    for col in sorted(parser.time_map.keys()):
        t = parser.time_map[col]
        if t != prev_time:
            cl = col_to_letter(col)
            print(f"  col{col:>3} ({cl:>3}列) -> {t}")
            prev_time = t
    SEP("-", 40)

    # 边框识别结果
    print()
    print(f"【边框识别】共 {len(parser.border_regions)} 个区域")
    SEP("-", 50)
    for i, r in enumerate(parser.border_regions, 1):
        sl = col_to_letter(r.start_col)
        el = col_to_letter(r.end_col)
        print(f"  {i}. 行{r.row}  {sl}{r.row}:{el}{r.row}"
              f"  列{r.start_col}-{r.end_col}  [{r.color}]")
    SEP("-", 50)

    # 架次信息
    print()
    if not sorties:
        print("  未解析到有效架次。")
    else:
        print(f"【架次信息】共 {len(sorties)} 个架次")
        for i, s in enumerate(sorties, 1):
            sl = col_to_letter(s.start_col)
            el = col_to_letter(s.end_col)
            print()
            SEP("-", 50)
            print(f"  架次 {i}")
            SEP("-", 50)
            print(f"  课目编号 : {s.course}")
            print(f"  起始时间 : {s.start_time}")
            print(f"  结束时间 : {s.end_time}")
            print(f"  所在行   : 第 {s.row} 行")
            print(f"  起始列   : 第 {s.start_col} 列 ({sl}列)")
            print(f"  结束列   : 第 {s.end_col} 列 ({el}列)")
            print(f"  边框颜色 : {s.border_color}")
            if s.front_seat_daizi:
                print(f"  前仓     : {s.front_seat_daizi}{s.front_seat_daihao}"
                      f"  (代字:{s.front_seat_daizi}  代号:{s.front_seat_daihao})")
            else:
                print(f"  前仓     : 未找到")
            if s.is_solo:
                print(f"  后仓     : 无 (单飞)")
            else:
                print(f"  后仓     : {s.rear_seat_daizi}{s.rear_seat_daihao}"
                      f"  (代字:{s.rear_seat_daizi}  代号:{s.rear_seat_daihao})")
            print(f"  飞行类型 : {'单飞 (仅前仓)' if s.is_solo else '带飞 (前仓+后仓)'}")

    print()
    SEP()
    print(f"  解析完成, 共 {len(sorties)} 个架次")
    SEP()


def get_filepath():
    if len(sys.argv) > 1:
        return sys.argv[1].strip().strip('"').strip("'")
    print()
    SEP()
    print("  飞行日计划表解析器 (边框识别版)")
    print("  请将Excel文件拖入此窗口, 然后按回车")
    SEP()
    print()
    path = input("文件路径: ").strip().strip('"').strip("'")
    return path


def main():
    filepath = get_filepath()

    if not filepath:
        print("错误: 未输入文件路径")
        input("\n按回车键退出...")
        sys.exit(1)

    if not os.path.isfile(filepath):
        print(f"错误: 文件不存在 -> {filepath}")
        input("\n按回车键退出...")
        sys.exit(1)

    try:
        parser = FlightPlanParser(filepath)
        sorties = parser.parse()
        print_results(parser, sorties, filepath)
    except Exception as e:
        print(f"\n解析失败: {e}")
        import traceback
        traceback.print_exc()

    input("\n按回车键退出...")


if __name__ == "__main__":
    main()
