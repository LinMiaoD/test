"""
Excel 加粗边框区域提取工具
功能: 扫描 Excel 文件，找出所有加粗边框的区域，输出：
  - 开始行 / 结束行
  - 开始列 / 结束列
  - 颜色分类（蓝色加粗 / 红色加粗 / 黑色加粗 / 其他）
  - 边框样式（thick / medium / double）

用法:
  python extract_bold_borders.py <excel文件路径> [sheet名称]
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from collections import defaultdict
import json, sys

BOLD_STYLES = {'thick', 'medium', 'double'}


def get_color_hex(color):
    if color is None:
        return None
    if color.type == 'rgb' and color.rgb and color.rgb != '00000000':
        return color.rgb
    if color.type == 'indexed':
        return f'indexed:{color.indexed}'
    if color.type == 'theme':
        return f'theme:{color.theme}'
    return None


def classify_color(hex_color):
    if hex_color is None:
        return 'black'
    h = hex_color.upper()
    # 处理 theme/indexed 类型
    if h.startswith('THEME:') or h.startswith('INDEXED:'):
        return h.lower()
    # 去掉 alpha 通道 (AARRGGBB -> RRGGBB)
    if len(h) == 8:
        h = h[2:]
    if len(h) != 6:
        return h
    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    if b > 150 and r < 100 and g < 100:
        return 'blue'
    if r > 150 and g < 100 and b < 100:
        return 'red'
    if r < 60 and g < 60 and b < 60:
        return 'black'
    if r < 100 and g > 150 and b < 100:
        return 'green'
    return f'rgb({r},{g},{b})'


def scan_bold_borders(filepath, sheet_name=None):
    wb = load_workbook(filepath, data_only=True)
    sheets = [sheet_name] if sheet_name else wb.sheetnames
    all_results = []

    for sname in sheets:
        ws = wb[sname]
        bold_cells = {}

        for row in ws.iter_rows():
            for cell in row:
                if cell.border is None:
                    continue
                info = {}
                for side_name in ('left', 'right', 'top', 'bottom'):
                    side = getattr(cell.border, side_name, None)
                    if side and side.style in BOLD_STYLES:
                        hex_c = get_color_hex(side.color)
                        info[side_name] = {
                            'style': side.style,
                            'color': classify_color(hex_c),
                            'hex': hex_c
                        }
                if info:
                    bold_cells[(cell.row, cell.column)] = info

        if not bold_cells:
            all_results.append({'sheet': sname, 'regions': [], 'message': '未发现加粗边框'})
            continue

        # 按 (颜色, 样式) 分组，计算各组的范围
        groups = defaultdict(list)
        for (r, c), sides in bold_cells.items():
            for side_name, detail in sides.items():
                key = (detail['color'], detail['style'])
                groups[key].append((r, c, side_name))

        regions = []
        for (color, style), cells_list in groups.items():
            rows = [x[0] for x in cells_list]
            cols = [x[1] for x in cells_list]
            sr, er = min(rows), max(rows)
            sc, ec = min(cols), max(cols)
            regions.append({
                'color': color,
                'color_label': f'{color}加粗' if color in ('blue','red','black','green') else f'{color} 加粗',
                'style': style,
                'start_row': sr,
                'end_row': er,
                'start_col': sc,
                'start_col_letter': get_column_letter(sc),
                'end_col': ec,
                'end_col_letter': get_column_letter(ec),
                'range': f'{get_column_letter(sc)}{sr}:{get_column_letter(ec)}{er}',
                'cell_count': len(set((r, c) for r, c, _ in cells_list))
            })

        all_results.append({'sheet': sname, 'regions': regions})

    wb.close()
    return all_results


def print_results(results):
    for sheet_data in results:
        print(f"\n{'='*50}")
        print(f"  Sheet: {sheet_data['sheet']}")
        print(f"{'='*50}")
        if not sheet_data['regions']:
            print("  未发现加粗边框")
            continue
        for i, region in enumerate(sheet_data['regions'], 1):
            print(f"\n  区域 {i}:")
            print(f"    类型:   {region['color_label']}（{region['style']}）")
            print(f"    范围:   {region['range']}")
            print(f"    行:     {region['start_row']} ~ {region['end_row']}")
            print(f"    列:     {region['start_col']}({region['start_col_letter']}) ~ {region['end_col']}({region['end_col_letter']})")
            print(f"    单元格: {region['cell_count']} 个")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("用法: python extract_bold_borders.py <excel文件> [sheet名称]")
        sys.exit(1)

    filepath = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else None
    results = scan_bold_borders(filepath, sheet)
    print_results(results)
    print("\n--- JSON ---")
    print(json.dumps(results, ensure_ascii=False, indent=2))
